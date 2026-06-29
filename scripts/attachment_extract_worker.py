from __future__ import annotations

import argparse
import base64
import csv
import json
from pathlib import Path


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("path")
    args = parser.parse_args()
    path = Path(args.path)
    suffix = path.suffix.lower()
    try:
        if suffix == ".pdf":
            text = _read_pdf(path)
            kind = "pdf"
        elif suffix in {".xlsx", ".xlsm"}:
            text = _read_xlsx(path)
            kind = "spreadsheet"
        elif suffix == ".csv":
            text = _read_csv(path)
            kind = "spreadsheet"
        else:
            text = ""
            kind = "file"
        payload = {
            "ok": True,
            "kind": kind,
            "text_b64": base64.b64encode(text.encode("utf-8")).decode("ascii"),
        }
    except Exception as exc:
        payload = {"ok": False, "kind": "file", "error": f"{type(exc).__name__}: {exc}"}
    print("ATTACHMENT_EXTRACT_JSON:" + json.dumps(payload, ensure_ascii=True))
    return 0


def _read_pdf(path: Path) -> str:
    try:
        from pdfminer.high_level import extract_text

        return extract_text(str(path)) or ""
    except Exception:
        from pypdf import PdfReader

        reader = PdfReader(str(path))
        return "\n".join(page.extract_text() or "" for page in reader.pages)


def _read_xlsx(path: Path) -> str:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    parts: list[str] = []
    for sheet in workbook.worksheets:
        parts.append(f"# Sheet: {sheet.title}")
        row_count = 0
        for row in sheet.iter_rows(values_only=True):
            values = [_format_cell(value) for value in row]
            if not any(values):
                continue
            parts.append("\t".join(values).rstrip())
            row_count += 1
            if row_count >= 80:
                parts.append("...")
                break
    return "\n".join(parts)


def _read_csv(path: Path) -> str:
    for encoding in ("utf-8-sig", "utf-8", "gb18030"):
        try:
            with path.open("r", encoding=encoding, newline="") as f:
                rows = []
                reader = csv.reader(f)
                for index, row in enumerate(reader):
                    if index >= 80:
                        rows.append(["..."])
                        break
                    rows.append(row)
                return "\n".join("\t".join(cell.strip() for cell in row) for row in rows)
        except UnicodeDecodeError:
            continue
    return path.read_text(encoding="utf-8", errors="replace")


def _format_cell(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


if __name__ == "__main__":
    raise SystemExit(main())
