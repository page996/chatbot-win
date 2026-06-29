from __future__ import annotations

import csv
import json
import shutil
from io import StringIO
from pathlib import Path
from typing import Any, Iterable


DEFAULT_TABLE_ROWS_PER_CHUNK = 100
SPREADSHEET_SUFFIXES = {".csv", ".xlsx", ".xlsm"}


def write_table_artifacts(
    source_path: str | Path,
    tables_dir: str | Path,
    *,
    rows_per_chunk: int = DEFAULT_TABLE_ROWS_PER_CHUNK,
) -> dict[str, Any]:
    """Parse a spreadsheet into JSON row chunks in the derived workspace."""

    source = Path(source_path)
    suffix = source.suffix.lower()
    output_dir = Path(tables_dir)
    if suffix not in SPREADSHEET_SUFFIXES:
        _reset_dir(output_dir)
        return {"status": "skipped", "reason": f"unsupported suffix: {suffix or 'unknown'}"}

    _reset_dir(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    index_path = output_dir / "index.json"
    try:
        if suffix == ".csv":
            chunks, tables = _parse_csv(source, output_dir, rows_per_chunk=rows_per_chunk)
        else:
            chunks, tables = _parse_xlsx(source, output_dir, rows_per_chunk=rows_per_chunk)
        row_count = sum(int(item.get("row_count", 0) or 0) for item in tables)
        payload = {
            "status": "completed" if chunks or tables else "empty",
            "source_path": str(source),
            "tables_dir": str(output_dir),
            "table_count": len(tables),
            "row_count": row_count,
            "chunk_count": len(chunks),
            "tables": tables,
            "chunks": chunks,
        }
    except Exception as exc:
        payload = {
            "status": "failed",
            "source_path": str(source),
            "tables_dir": str(output_dir),
            "table_count": 0,
            "row_count": 0,
            "chunk_count": 0,
            "tables": [],
            "chunks": [],
            "error": f"{type(exc).__name__}: {exc}",
        }

    _write_json(index_path, payload)
    payload["index_path"] = str(index_path)
    return payload


def _parse_csv(source: Path, output_dir: Path, *, rows_per_chunk: int) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    chunks: list[dict[str, Any]] = []
    tables: list[dict[str, Any]] = []
    reader = csv.reader(StringIO(_read_csv_text(source)))
    headers: list[str] | None = None
    chunk_rows: list[dict[str, Any]] = []
    chunk_start = 0
    row_count = 0
    source_row_number = 0
    for raw_row in reader:
        source_row_number += 1
        row = [_cell_to_json_value(cell) for cell in raw_row]
        if not _row_has_value(row):
            continue
        if headers is None:
            headers = _unique_headers([str(cell) for cell in row])
            continue
        row_count += 1
        if not chunk_rows:
            chunk_start = source_row_number
        chunk_rows.append(_row_to_object(headers, row))
        if len(chunk_rows) >= rows_per_chunk:
            chunks.append(
                _write_chunk(
                    output_dir,
                    table_index=1,
                    chunk_index=len(chunks) + 1,
                    sheet_name=source.stem,
                    columns=headers,
                    rows=chunk_rows,
                    source_start_row=chunk_start,
                    source_end_row=source_row_number,
                )
            )
            chunk_rows = []
    if headers is not None and chunk_rows:
        chunks.append(
            _write_chunk(
                output_dir,
                table_index=1,
                chunk_index=len(chunks) + 1,
                sheet_name=source.stem,
                columns=headers,
                rows=chunk_rows,
                source_start_row=chunk_start,
                source_end_row=source_row_number,
            )
        )
    if headers is not None:
        tables.append(
            {
                "table_index": 1,
                "sheet_name": source.stem,
                "columns": headers,
                "row_count": row_count,
                "chunk_count": len(chunks),
            }
        )
    return chunks, tables


def _parse_xlsx(source: Path, output_dir: Path, *, rows_per_chunk: int) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - depends on local optional dependency
        raise RuntimeError("openpyxl is required for xlsx/xlsm table artifacts") from exc

    workbook = load_workbook(source, read_only=True, data_only=True)
    chunks: list[dict[str, Any]] = []
    tables: list[dict[str, Any]] = []
    for table_index, sheet in enumerate(workbook.worksheets, start=1):
        headers: list[str] | None = None
        chunk_rows: list[dict[str, Any]] = []
        chunk_start = 0
        row_count = 0
        first_chunk_index = len(chunks) + 1
        for source_row_number, raw_row in enumerate(sheet.iter_rows(values_only=True), start=1):
            row = [_cell_to_json_value(cell) for cell in raw_row]
            if not _row_has_value(row):
                continue
            if headers is None:
                headers = _unique_headers([str(cell) for cell in row])
                continue
            row_count += 1
            if not chunk_rows:
                chunk_start = source_row_number
            chunk_rows.append(_row_to_object(headers, row))
            if len(chunk_rows) >= rows_per_chunk:
                chunks.append(
                    _write_chunk(
                        output_dir,
                        table_index=table_index,
                        chunk_index=len(chunks) + 1,
                        sheet_name=sheet.title,
                        columns=headers,
                        rows=chunk_rows,
                        source_start_row=chunk_start,
                        source_end_row=source_row_number,
                    )
                )
                chunk_rows = []
        if headers is not None and chunk_rows:
            chunks.append(
                _write_chunk(
                    output_dir,
                    table_index=table_index,
                    chunk_index=len(chunks) + 1,
                    sheet_name=sheet.title,
                    columns=headers,
                    rows=chunk_rows,
                    source_start_row=chunk_start,
                    source_end_row=source_row_number,
                )
            )
        if headers is not None:
            tables.append(
                {
                    "table_index": table_index,
                    "sheet_name": sheet.title,
                    "columns": headers,
                    "row_count": row_count,
                    "chunk_count": len(chunks) - first_chunk_index + 1 if row_count else 0,
                }
            )
    workbook.close()
    return chunks, tables


def _write_chunk(
    output_dir: Path,
    *,
    table_index: int,
    chunk_index: int,
    sheet_name: str,
    columns: list[str],
    rows: list[dict[str, Any]],
    source_start_row: int,
    source_end_row: int,
) -> dict[str, Any]:
    path = output_dir / f"table_{table_index:04d}_chunk_{chunk_index:04d}.json"
    payload = {
        "table_index": table_index,
        "chunk_index": chunk_index,
        "sheet_name": sheet_name,
        "source_range": {"start_row": source_start_row, "end_row": source_end_row},
        "columns": columns,
        "row_count": len(rows),
        "rows": rows,
    }
    _write_json(path, payload)
    return {
        "table_index": table_index,
        "chunk_index": chunk_index,
        "sheet_name": sheet_name,
        "path": str(path),
        "source_range": payload["source_range"],
        "columns": columns,
        "row_count": len(rows),
    }


def _read_csv_text(source: Path) -> str:
    content = source.read_bytes()
    for encoding in ("utf-8-sig", "utf-8", "gb18030"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    return content.decode("utf-8", errors="replace")


def _unique_headers(values: Iterable[str]) -> list[str]:
    seen: dict[str, int] = {}
    headers: list[str] = []
    for index, value in enumerate(values, start=1):
        base = str(value).strip() or f"column_{index}"
        count = seen.get(base, 0) + 1
        seen[base] = count
        headers.append(base if count == 1 else f"{base}_{count}")
    return headers


def _row_to_object(headers: list[str], row: list[Any]) -> dict[str, Any]:
    result: dict[str, Any] = {}
    for index, value in enumerate(row, start=1):
        key = headers[index - 1] if index <= len(headers) else f"column_{index}"
        result[key] = value
    for index in range(len(row) + 1, len(headers) + 1):
        result[headers[index - 1]] = ""
    return result


def _cell_to_json_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, (bool, int, float)):
        return value
    isoformat = getattr(value, "isoformat", None)
    if callable(isoformat):
        return isoformat()
    return str(value).strip()


def _row_has_value(row: list[Any]) -> bool:
    return any(str(cell).strip() for cell in row)


def _reset_dir(path: Path) -> None:
    if path.exists():
        if path.name != "tables":
            raise PermissionError(f"refusing to reset non-table artifact directory: {path}")
        shutil.rmtree(path)


def _write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_suffix(path.suffix + ".tmp")
    with tmp.open("w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
    tmp.replace(path)
